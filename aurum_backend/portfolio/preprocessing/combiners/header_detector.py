#!/usr/bin/env python3
"""
Header Detection Utility

Dynamically detects header rows in Excel files by scanning for expected column patterns.
Used for banks like CS where header positions may vary.
"""

import logging
from pathlib import Path
from typing import List, Optional
import pandas as pd

logger = logging.getLogger(__name__)


class HeaderDetector:
    """Utility class for dynamically detecting header rows in Excel files."""
    
    # Expected column patterns for different file types
    SECURITIES_KEY_COLUMNS = [
        'Asset Category',
        'Asset Subcategory', 
        'Currency (for Nominal Field only)',
        'Nominal/Number',
        'Description',
        'Valor',
        'Maturity',
        'Currency (Price)',
        'Price'
    ]
    
    # Pershing-specific column patterns
    PERSHING_SECURITIES_KEY_COLUMNS = [
        'Security ID',
        'CUSIP',
        'Account Number',
        'Account Nickname/Title',
        'Description',
        'Asset Classification',
        'Quantity',
        'Price'
    ]
    
    PERSHING_UNITCOST_KEY_COLUMNS = [
        'Security',
        'Account Number',
        'Account Nickname',
        'Quantity',
        'Unit Cost',
        'Last Price',
        'Total Cost',
        'Market Value'
    ]
    
    PERSHING_TRANSACTIONS_KEY_COLUMNS = [
        'Date',
        'Type',
        'Activity Description',
        'Net Amount',
        'Details',
        'Trade Date',
        'Quantity',
        'Price'
    ]
    
    TRANSACTIONS_KEY_COLUMNS = [
        'Booking Date',
        'Text',
        'Debit',
        'Credit',
        'Value Date',
        'Balance',
        'Cantidad',
        'ID',
        'Precio'
    ]
    
    # HSBC-specific column patterns (same as Pershing since format is identical)
    HSBC_SECURITIES_KEY_COLUMNS = [
        'Security ID',
        'CUSIP',
        'Account Number',
        'Account Nickname/Title',
        'Description',
        'Asset Classification',
        'Quantity',
        'Price'
    ]
    
    HSBC_UNITCOST_KEY_COLUMNS = [
        'Security',
        'Account Number',
        'Account Nickname',
        'Quantity',
        'Unit Cost',
        'Last Price',
        'Total Cost',
        'Market Value'
    ]
    
    HSBC_TRANSACTIONS_KEY_COLUMNS = [
        'Date',
        'Account',
        'Type',
        'Activity Description',
        'Net Amount',
        'Details',
        'Trade Date',
        'Quantity',
        'Price'
    ]
    
    @classmethod
    def find_securities_header_row(cls, file_path: Path, max_search_rows: int = 15) -> int:
        """
        Find header row for securities files by scanning for expected column patterns.
        
        Args:
            file_path: Path to Excel file
            max_search_rows: Maximum rows to search (default: 15)
            
        Returns:
            Row index (0-based) where headers are found
            
        Raises:
            ValueError: If headers not found within search range
        """
        return cls._find_header_row(
            file_path=file_path,
            expected_columns=cls.SECURITIES_KEY_COLUMNS,
            file_type='securities',
            max_search_rows=max_search_rows
        )
    
    @classmethod
    def find_transactions_header_row(cls, file_path: Path, max_search_rows: int = 15) -> int:
        """
        Find header row for transactions files by scanning for expected column patterns.
        
        Args:
            file_path: Path to Excel file
            max_search_rows: Maximum rows to search (default: 15)
            
        Returns:
            Row index (0-based) where headers are found
            
        Raises:
            ValueError: If headers not found within search range
        """
        return cls._find_header_row(
            file_path=file_path,
            expected_columns=cls.TRANSACTIONS_KEY_COLUMNS,
            file_type='transactions',
            max_search_rows=max_search_rows
        )
    
    @classmethod
    def find_pershing_securities_header_row(cls, file_path: Path, max_search_rows: int = 15) -> int:
        """
        Find header row for Pershing securities files.
        
        Args:
            file_path: Path to Excel file
            max_search_rows: Maximum rows to search (default: 15)
            
        Returns:
            Row index (0-based) where headers are found
            
        Raises:
            ValueError: If headers not found within search range
        """
        return cls._find_header_row(
            file_path=file_path,
            expected_columns=cls.PERSHING_SECURITIES_KEY_COLUMNS,
            file_type='pershing_securities',
            max_search_rows=max_search_rows
        )
    
    @classmethod
    def find_pershing_unitcost_header_row(cls, file_path: Path, max_search_rows: int = 15) -> int:
        """
        Find header row for Pershing unitcost files.
        
        Args:
            file_path: Path to Excel file
            max_search_rows: Maximum rows to search (default: 15)
            
        Returns:
            Row index (0-based) where headers are found
            
        Raises:
            ValueError: If headers not found within search range
        """
        return cls._find_header_row(
            file_path=file_path,
            expected_columns=cls.PERSHING_UNITCOST_KEY_COLUMNS,
            file_type='pershing_unitcost',
            max_search_rows=max_search_rows
        )
    
    @classmethod
    def find_pershing_transactions_header_row(cls, file_path: Path, max_search_rows: int = 15) -> int:
        """
        Find header row for Pershing transactions files.
        
        Args:
            file_path: Path to Excel file
            max_search_rows: Maximum rows to search (default: 15)
            
        Returns:
            Row index (0-based) where headers are found
            
        Raises:
            ValueError: If headers not found within search range
        """
        return cls._find_header_row(
            file_path=file_path,
            expected_columns=cls.PERSHING_TRANSACTIONS_KEY_COLUMNS,
            file_type='pershing_transactions',
            max_search_rows=max_search_rows
        )
    
    @classmethod
    def find_hsbc_securities_header_row(cls, file_path: Path, max_search_rows: int = 20) -> int:
        """
        Find header row for HSBC securities files.
        
        Args:
            file_path: Path to Excel file
            max_search_rows: Maximum rows to search (default: 20)
            
        Returns:
            Row index (0-based) where headers are found
            
        Raises:
            ValueError: If headers not found within search range
        """
        return cls._find_header_row(
            file_path=file_path,
            expected_columns=cls.HSBC_SECURITIES_KEY_COLUMNS,
            file_type='hsbc_securities',
            max_search_rows=max_search_rows
        )
    
    @classmethod
    def find_hsbc_unitcost_header_row(cls, file_path: Path, max_search_rows: int = 20) -> int:
        """
        Find header row for HSBC unitcost files.
        
        Args:
            file_path: Path to Excel file
            max_search_rows: Maximum rows to search (default: 20)
            
        Returns:
            Row index (0-based) where headers are found
            
        Raises:
            ValueError: If headers not found within search range
        """
        return cls._find_header_row(
            file_path=file_path,
            expected_columns=cls.HSBC_UNITCOST_KEY_COLUMNS,
            file_type='hsbc_unitcost',
            max_search_rows=max_search_rows
        )
    
    @classmethod
    def find_hsbc_transactions_header_row(cls, file_path: Path, max_search_rows: int = 20) -> int:
        """
        Find header row for HSBC transactions files.
        
        Args:
            file_path: Path to Excel file
            max_search_rows: Maximum rows to search (default: 20)
            
        Returns:
            Row index (0-based) where headers are found
            
        Raises:
            ValueError: If headers not found within search range
        """
        return cls._find_header_row(
            file_path=file_path,
            expected_columns=cls.HSBC_TRANSACTIONS_KEY_COLUMNS,
            file_type='hsbc_transactions',
            max_search_rows=max_search_rows
        )
    
    @classmethod
    def _find_header_row(cls, file_path: Path, expected_columns: List[str], 
                        file_type: str, max_search_rows: int = 15) -> int:
        """
        Internal method to find header row by scanning for expected column patterns.
        
        Args:
            file_path: Path to Excel file
            expected_columns: List of column names to search for
            file_type: Type of file ('securities' or 'transactions')
            max_search_rows: Maximum rows to search
            
        Returns:
            Row index (0-based) where headers are found
            
        Raises:
            ValueError: If headers not found within search range
        """
        logger.debug(f"🔍 Scanning for {file_type} headers in: {file_path.name}")
        
        try:
            # Read first N rows to search for headers (no header specified)
            preview_df = pd.read_excel(file_path, nrows=max_search_rows, header=None)
            
            if preview_df.empty:
                raise ValueError(f"File is empty: {file_path}")
            
            logger.debug(f"  📊 Scanning {len(preview_df)} rows for header patterns...")
            
            # Search each row for header patterns
            for row_idx in range(len(preview_df)):
                row_values = preview_df.iloc[row_idx].astype(str).str.strip()
                
                # Count matches (case-insensitive, partial matching)
                matches = 0
                matched_columns = []
                
                for expected_col in expected_columns:
                    for cell_value in row_values:
                        if pd.notna(cell_value) and expected_col.lower() in cell_value.lower():
                            matches += 1
                            matched_columns.append(expected_col)
                            break
                
                # Calculate match percentage
                match_percentage = matches / len(expected_columns)
                
                logger.debug(f"  Row {row_idx + 1}: {matches}/{len(expected_columns)} columns matched ({match_percentage:.1%})")
                
                # If we find 4/9 (44%+) of expected columns, this is likely the header row
                if matches >= 4:  # 4/9 columns as approved
                    logger.info(f"✅ Found {file_type} headers at row {row_idx + 1}")
                    logger.info(f"  📋 Matched columns: {', '.join(matched_columns)}")
                    logger.info(f"  📊 Match rate: {matches}/{len(expected_columns)} ({match_percentage:.1%})")
                    return row_idx
            
            # If we get here, no suitable header row was found
            logger.warning(f"⚠️ Could not find {file_type} header row in first {max_search_rows} rows of {file_path.name}")
            raise ValueError(f"Could not find {file_type} header row in {file_path}")
            
        except Exception as e:
            logger.error(f"❌ Error scanning for headers in {file_path}: {str(e)}")
            raise
    
    @classmethod
    def validate_columns(cls, df: pd.DataFrame, expected_columns: List[str], 
                        min_match_ratio: float = 0.4) -> bool:
        """
        Validate that a DataFrame contains expected columns.
        
        Args:
            df: DataFrame to validate
            expected_columns: List of expected column names
            min_match_ratio: Minimum ratio of columns that must match
            
        Returns:
            True if validation passes, False otherwise
        """
        if df.empty:
            return False
        
        # Get actual column names (convert to string and strip whitespace)
        actual_columns = [str(col).strip().lower() for col in df.columns if pd.notna(col)]
        
        # Count matches
        matches = 0
        for expected_col in expected_columns:
            if any(expected_col.lower() in actual_col for actual_col in actual_columns):
                matches += 1
        
        match_ratio = matches / len(expected_columns)
        
        logger.debug(f"Column validation: {matches}/{len(expected_columns)} matched ({match_ratio:.1%})")
        
        return match_ratio >= min_match_ratio
    
    @classmethod
    def read_excel_with_fallback(cls, file_path: Path, header_row: int = 0) -> pd.DataFrame:
        """
        Read Excel file with support for both .xlsx and .xls formats.
        
        Args:
            file_path: Path to Excel file
            header_row: Row index for headers
            
        Returns:
            DataFrame
            
        Raises:
            Exception: If file cannot be read with any engine
        """
        logger.debug(f"📖 Reading Excel file: {file_path.name} (header row: {header_row})")
        
        # Try different engines based on file extension
        engines_to_try = []
        
        if file_path.suffix.lower() == '.xlsx':
            engines_to_try = ['openpyxl', 'xlrd']
        elif file_path.suffix.lower() == '.xls':
            engines_to_try = ['xlrd', 'openpyxl']
        else:
            engines_to_try = ['openpyxl', 'xlrd']
        
        last_error = None
        
        for engine in engines_to_try:
            try:
                df = pd.read_excel(file_path, header=header_row, engine=engine)
                logger.debug(f"  ✅ Successfully read with {engine} engine")
                return df
            except Exception as e:
                logger.debug(f"  ❌ Failed with {engine} engine: {str(e)}")
                last_error = e
                continue
        
        # If we get here, all engines failed
        logger.error(f"❌ Failed to read {file_path} with any engine")
        raise last_error or Exception(f"Could not read {file_path}")

    @classmethod
    def read_excel_with_outline_awareness(cls, file_path: Path, header_row: int, 
                                        filter_summary_only: bool = False) -> pd.DataFrame:
        """
        Read Excel file with awareness of Excel grouping/outlining.
        
        For Pershing unitcost files that use Excel grouping, this can filter to only 
        read summary rows (outline_level=0, not hidden) which contain the totals
        that are visible by default in Excel.
        
        Args:
            file_path: Path to Excel file
            header_row: Row index (0-based) where headers are located
            filter_summary_only: If True, only read summary rows (for unitcost files)
            
        Returns:
            DataFrame with file data
        """
        try:
            if not filter_summary_only:
                # Standard reading
                return cls.read_excel_with_fallback(file_path, header_row)
            
            # Outline-aware reading for unitcost files
            from openpyxl import load_workbook
            
            logger.debug(f"  📊 Reading Excel with outline awareness: {file_path.name}")
            
            # Load workbook to check outline levels
            wb = load_workbook(file_path)
            ws = wb.active
            
            # Find rows that are summary rows (outline_level=0 and not hidden)
            summary_row_indices = []
            data_start = header_row + 1
            
            for row_num in range(data_start, data_start + 200):  # Check reasonable range
                # Check if row has data in Security column (usually column 2)
                cell_value = ws.cell(row_num, 2).value
                if cell_value is None or str(cell_value).strip() == '':
                    break  # End of data
                
                outline_level = getattr(ws.row_dimensions[row_num], 'outline_level', 0)
                hidden = getattr(ws.row_dimensions[row_num], 'hidden', False)
                
                if outline_level == 0 and not hidden:
                    # This is a summary row (1-based for Excel, 0-based for pandas)
                    summary_row_indices.append(row_num - 1)  # Convert to 0-based
            
            logger.debug(f"  📋 Found {len(summary_row_indices)} summary rows")
            
            if not summary_row_indices:
                logger.warning(f"  ⚠️ No summary rows found, falling back to standard reading")
                return cls.read_excel_with_fallback(file_path, header_row)
            
            # Read the entire file first
            df_all = cls.read_excel_with_fallback(file_path, header_row)
            
            if df_all.empty:
                return df_all
            
            # Filter to only summary rows
            # summary_row_indices are relative to start of file, need to adjust for header
            data_indices = [idx - header_row - 1 for idx in summary_row_indices if idx > header_row]
            data_indices = [idx for idx in data_indices if 0 <= idx < len(df_all)]
            
            if not data_indices:
                logger.warning(f"  ⚠️ No valid data indices after filtering")
                return df_all
            
            df_summary = df_all.iloc[data_indices].copy()
            logger.info(f"  ✅ Filtered to {len(df_summary)} summary rows (from {len(df_all)} total)")
            
            return df_summary
            
        except Exception as e:
            logger.error(f"  ❌ Error in outline-aware reading: {str(e)}")
            logger.warning(f"  ⚠️ Falling back to standard Excel reading")
            return cls.read_excel_with_fallback(file_path, header_row) 